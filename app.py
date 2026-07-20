from flask import Flask, render_template, jsonify, request, send_file
import json
import logging
from ddgs import DDGS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io
from datetime import datetime

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)

# Загрузка данных
with open('data/equipment.json', 'r', encoding='utf-8') as f:
    equipment = json.load(f)

with open('data/spares.json', 'r', encoding='utf-8') as f:
    spares = json.load(f)

COMMERCIAL_KEYWORDS = ['купить', 'цена', 'руб', 'стоимость', 'завод', 'ТД', 'торговый дом', 'каталог', 'прайс', 'магазин', 'предложение']

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/equipment')
def get_equipment():
    return jsonify(equipment)

@app.route('/api/spares')
def get_spares():
    return jsonify(spares)

@app.route('/api/search-analogs', methods=['POST'])
def search_analogs():
    data = request.get_json()
    name = data.get('name', '')
    part_number = data.get('part_number', '')
    manufacturer = data.get('manufacturer', '')

    query_parts = [name, "купить", "цена"]
    if manufacturer and manufacturer.lower() not in name.lower():
        query_parts.append(manufacturer)
    if part_number and part_number not in ["6305-2RS1", "2L5P-2RS1"]:
        query_parts.append(part_number)
    query = " ".join(query_parts)
    app.logger.info(f"Поисковый запрос: {query}")

    try:
        with DDGS(timeout=30) as ddgs:
            results = list(ddgs.text(query, max_results=10, backend='html'))
            app.logger.info(f"Найдено до фильтрации: {len(results)}")
            filtered = []
            for item in results:
                title = item.get('title', '').lower()
                body = item.get('body', '').lower()
                combined = title + ' ' + body
                if any(kw in combined for kw in COMMERCIAL_KEYWORDS):
                    filtered.append({
                        'title': item.get('title', 'Без названия'),
                        'url': item.get('href', '#'),
                        'description': item.get('body', '')
                    })
            final = filtered[:5]
            app.logger.info(f"После фильтрации: {len(final)}")
            return jsonify({'success': True, 'results': final})
    except Exception as e:
        app.logger.error(f"Ошибка поиска: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

# ------------------- ЭКСПОРТ В EXCEL -------------------
@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    results = data.get('results', [])
    spare_name = data.get('spare_name', 'Неизвестная запчасть')

    wb = Workbook()
    ws = wb.active
    ws.title = "Аналоги"

    # Заголовки
    headers = ['№', 'Производитель', 'Наименование', 'Цена', 'Телефон', 'Адрес', 'Сайт']
    ws.append(headers)

    # Стиль заголовка
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2a5a8a", end_color="2a5a8a", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Заполнение строк
    for idx, item in enumerate(results, start=1):
        title = item.get('title', '')
        manufacturer = '—'
        if 'завод' in title.lower():
            manufacturer = 'Завод (уточнить)'
        elif 'ооо' in title.lower():
            manufacturer = 'ООО (уточнить)'
        elif 'торговый дом' in title.lower():
            manufacturer = 'ТД (уточнить)'

        ws.append([
            idx,
            manufacturer,
            spare_name,
            '—',
            '—',
            '—',
            item.get('url', '#')
        ])

    # Автоширина колонок
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        adjusted = min(max_len + 2, 40)
        ws.column_dimensions[col[0].column_letter].width = adjusted

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"analogs_{spare_name[:20]}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ------------------- ЭКСПОРТ В PDF (исправленный) -------------------
@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    data = request.get_json()
    results = data.get('results', [])
    spare_name = data.get('spare_name', 'Неизвестная запчасть')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=12,
        textColor=colors.HexColor('#1a2634')
    )
    normal_style = styles['Normal']

    story = []

    # Заголовок отчета
    story.append(Paragraph(f"Отчёт по аналогам: {spare_name}", title_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Дата генерации: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 0.5*cm))

    # Подготовка данных таблицы
    table_data = [['№', 'Производитель', 'Наименование', 'Цена', 'Телефон', 'Адрес', 'Сайт']]
    for idx, item in enumerate(results, start=1):
        title = item.get('title', '')
        manufacturer = '—'
        if 'завод' in title.lower():
            manufacturer = 'Завод (уточнить)'
        elif 'ооо' in title.lower():
            manufacturer = 'ООО (уточнить)'
        elif 'торговый дом' in title.lower():
            manufacturer = 'ТД (уточнить)'

        # Ограничиваем длину текста, чтобы таблица не разрывалась
        site = item.get('url', '#')
        if len(site) > 40:
            site = site[:37] + '...'

        table_data.append([
            str(idx),
            manufacturer,
            spare_name,
            '—',
            '—',
            '—',
            site
        ])

    # Создание таблицы с шириной колонок
    col_widths = [1.2*cm, 3.5*cm, 4*cm, 2.5*cm, 2.5*cm, 3.5*cm, 4.5*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2a5a8a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('WORDWRAP', (0,0), (-1,-1), True),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))

    story.append(table)
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("Источник: поиск в интернете (DuckDuckGo)", normal_style))
    story.append(Paragraph("Данные о ценах и контактах требуют уточнения у поставщиков.", normal_style))

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"analogs_{spare_name[:20]}_{datetime.now().strftime('%Y%m%d')}.pdf",
        mimetype='application/pdf'
    )

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)